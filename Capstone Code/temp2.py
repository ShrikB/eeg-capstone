import openpyxl
from decimal import Decimal, InvalidOperation
from scipy import signal
import numpy as np
import matplotlib.pyplot as plt
from scipy.fft import fft
from scipy.stats import zscore

def search_excel_file(filename, search_string):
    # Load the excel file
    wb = openpyxl.load_workbook(filename)

    # Select the active worksheet
    ws = wb.active

    # Search for the cell containing the search string
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == search_string:
                # Get the row and column of the cell
                col_num = cell.column
                row_num = cell.row

                # Return the row and column of the cell
                return (row_num, col_num)

    # If the search string was not found, return None
    return False 

def get_next_N_values(filename, row_num, col_num, N):
    # Load the excel file
    wb = openpyxl.load_workbook(filename)

    # Select the active worksheet
    ws = wb.active

    # Get the starting cell
    start_cell = ws.cell(row=row_num, column=col_num)

    # Get the values of the next N cells below the starting cell
    values = []
    for i in range(N):
        cell = ws.cell(row=start_cell.row+i+1, column=start_cell.column)
        value = cell.value
        values.append(value)

    # Return the values as an array
    return values

def convert_to_decimal(lst):
    result = []
    for item in lst:
        try:
            # Try to convert the string to a Decimal
            num = Decimal(item)
            # Round the Decimal to the 15th significant figure
            num = round(num, 14 - num.adjusted())
            # Append the rounded Decimal to the result list
            result.append(float(num))
            print("appended")
        except InvalidOperation:
            # If the string cannot be converted to a Decimal, skip it
            pass
    return result

def notch_filter(data, fs, target_freq, Q):
    # Calculate the notch filter frequency and bandwidth
    w0 = target_freq / (fs / 2)
    bw = w0 / Q
    # Create the filter coefficients for a notch filter
    b, a = signal.iirnotch(w0, bw, fs=fs)
    # Apply the filter to the data
    filtered_data = signal.filtfilt(b, a, data)
    return filtered_data

def lowpass_filter(data, fs, cutoff_freq, filter_order):
    # Calculate the filter coefficients for a lowpass filter
    nyquist_freq = fs / 2
    normalized_cutoff_freq = cutoff_freq / nyquist_freq
    b, a = signal.butter(filter_order, normalized_cutoff_freq, btype='lowpass')
    # Apply the filter to the data
    filtered_data = signal.filtfilt(b, a, data)
    return filtered_data

def highpass_filter(data, fs, cutoff_freq, filter_order):
    # Calculate the filter coefficients for a highpass filter
    nyquist_freq = fs / 2
    normalized_cutoff_freq = cutoff_freq / nyquist_freq
    b, a = signal.butter(filter_order, normalized_cutoff_freq, btype='highpass')
    # Apply the filter to the data
    filtered_data = signal.filtfilt(b, a, data)
    return filtered_data

def plot_data(data):
    # Remove outliers
    data = data[(np.abs((data - np.mean(data))) / np.std(data)) < 3]
    
    # Plot the data
    plt.plot(data)
    plt.xlabel('Data Point')
    plt.ylabel('Value')
    plt.show()

# def plot_fft(data_list, fs, limit_freq=50, z_thresh=3):
#     fig, axs = plt.subplots(len(data_list), 2, figsize=(20, len(data_list)*7))

#     for i, data in enumerate(data_list):

#         # plot title
#         outliers = group_outliers(data)

#         data_clean = data[(np.abs(zscore(data)) < z_thresh)]
#         N = len(data_clean)
#         T = 1 / fs
#         yf = fft(data_clean)
#         xf = np.linspace(0.0, 1.0/(2.0*T), N//2)
#         yf = 2.0/N * np.abs(yf[0:N//2])

#         axs[i, 0].plot(data_clean)
#         # axs[i, 0].set_xlabel('Time')
#         # axs[i, 0].set_ylabel('Amplitude')

#         # title = f'Time domain plot {i+1} ' + str(len(outliers))  + ' ' + outliers

#         title = f'Time domain plot {i+1} REM cycles: {len(outliers)}'


#         # axs[i, 0].set_title(f'Time domain plot {i+1}')
#         axs[i, 0].set_title(title)

#         axs[i, 1].plot(xf, yf)
#         # axs[i, 1].set_xlabel('Frequency')
#         # axs[i, 1].set_ylabel('Amplitude')
#         axs[i, 1].set_title(f'FFT plot {i+1}')
#         axs[i, 1].set_xlim([0, limit_freq])
#         # axs[i, 1].xticks(np.arange(0, 50, 1.0))

#     fig.tight_layout(pad=12)
#     plt.show()

def plot_fft(data_list, fs, limit_freq=50, z_thresh=3):
    fig, axs = plt.subplots(len(data_list), 2, figsize=(20, len(data_list)*7))

    for i, data in enumerate(data_list):

        # plot title 
        outliers, sp_ep_pairs = group_outliers(data)

        data_clean = data[(np.abs(zscore(data)) < z_thresh)]
        N = len(data_clean)
        T = 1 / fs
        yf = fft(data_clean)
        xf = np.linspace(0.0, 1.0/(2.0*T), N//2)
        yf = 2.0/N * np.abs(yf[0:N//2])

        axs[i, 0].plot(data_clean)

        # add vertical red lines at positions indicated by sp_ep_pairs
        for sp_ep in sp_ep_pairs:
            axs[i, 0].axvline(x=sp_ep[0], color='red')
            axs[i, 0].axvline(x=sp_ep[1], color='red')

        title = f'Time domain plot {i+1} REM cycles: {len(outliers)}'

        axs[i, 0].set_title(title)

        axs[i, 1].plot(xf, yf)
        axs[i, 1].set_title(f'FFT plot {i+1}')
        axs[i, 1].set_xlim([0, limit_freq])
        axs[i, 1].set_xticks(np.arange(0, limit_freq+1, 5))

    fig.tight_layout(pad=12)
    plt.show()

def filter_zero_lists(lst):
    return [sublst for sublst in lst if any(sublst)]

def group_outliers(background_noise, threshold=2.5, window_size=8000):
    # ignore the first 7000 data points
    background_noise = background_noise[7000:]

    # calculate the mean of the background noise
    noise_mean = np.mean(background_noise)

    # calculate the threshold for identifying outliers
    noise_threshold = threshold * np.std(background_noise)

    # create an array to store the outlier indices
    outlier_indices = []

    # scan the background noise and identify the outliers
    for i in range(len(background_noise)):
        if abs(background_noise[i] - noise_mean) > noise_threshold:
            outlier_indices.append(i)

    # group the outlier indices into sets
    outlier_sets = []
    current_set = []
    for i in range(len(outlier_indices)):
        if i == 0:
            current_set.append(outlier_indices[i])
        elif outlier_indices[i] - outlier_indices[i-1] <= window_size:
            current_set.append(outlier_indices[i])
        else:
            outlier_sets.append(current_set)
            current_set = [outlier_indices[i]]

    # add the last set to the list of outlier sets
    outlier_sets.append(current_set)

    # print the starting index of each outlier set and the number of sets
    print("Number of outlier sets:", len(outlier_sets))

    sp_ep_pairs = []

    for outlier_set in outlier_sets:
        print("Starting index:", outlier_set[0] + 7000)
        sp = outlier_set[0] + 7000
        ep = outlier_set[-1] + 7000
        sp_ep_pairs.append((sp,ep))

    return outlier_sets, sp_ep_pairs





sheetname = "shrikantData.xlsx"
searchString = ["CH2", "CH3", "CH4", "CH5", "CH6", "CH7", "CH8"]
# searchString = ["CH2", "CH3"]
channels = []
channelData = []
filteredData = []

for word in searchString:
    rv = search_excel_file(sheetname, word)
    if(rv != False):
        row, column = rv
        channels.append({"CH": word, "ROW": row, "COLUMN": column})

for channel in channels:
    channelData.append(get_next_N_values(sheetname, channel["ROW"], channel["COLUMN"], 89900))

channelData = filter_zero_lists(channelData)

for data in channelData:
    x = data
    x = notch_filter(x, 500, 60, 2)
    x = lowpass_filter(x, 500, 60, 3)
    x = highpass_filter(x, 500, 0.1, 3)
    filteredData.append(x)

# group_outliers(filteredData[0])
# group_outliers(filteredData[1])
# group_outliers(filteredData[2])
# group_outliers(filteredData[3])

plot_fft(filteredData, 500)


