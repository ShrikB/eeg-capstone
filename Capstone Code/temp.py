import numpy
import pandas
from openpyxl import load_workbook

# wb = load_workbook("chrisData.xlsx")
# ws = wb.active

# # c = ws['A38']
# c = ws.cell(38,1)

# print(c)
        
        
# for rows in ws.rows:
#     for cells in rows:


df = pandas.read_excel('chrisData.xlsx')

a = df.where(df=='CH8')

print(a.index)


